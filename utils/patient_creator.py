import os
import json
import openpyxl
from dataclasses import dataclass
from openpyxl.utils import coordinate_to_tuple
from datetime import datetime, timedelta
from docx2python.utilities import replace_docx_text
from random import randint, uniform
from utils.GSheet import spreadsheet_fill


@dataclass
class Patient:
    """
    Basic class for patient
    """
    name: str
    birthday: str
    history: str
    department: str
    hospitalisation_date: str
    diagnosis_main: str
    operation: str
    operation_date: str
    diagnosis_secondary: str
    icu_date: str
    doctor: str
    path: str

    def __post_init__(self):
        """
        Calculates age attr
        :return:
        """
        birthday_dt = datetime.strptime(self.birthday, "%d.%m.%Y")
        today = datetime.today()

        self.age = str(
            today.year - birthday_dt.year - ((today.month, today.day) < (birthday_dt.month, birthday_dt.day)))

    def create_replacements(self):
        """
        Creates generalized replacements dict
        :return: dict of key_word: replacement
        """
        return {
            "{NAME}": self.name,
            "{AGE}": self.age,
            "{BIRTHDAY}": self.birthday,
            "{HISTORY}": self.history,
            "{DEPARTMENT}": self.department,
            "{HOSPITALDATE}": self.hospitalisation_date,
            "{DXMAIN}": self.diagnosis_main,
            "{OPERATION}": self.operation,
            "{OPERATIONDATE}": self.operation_date,
            "{DXSECONDARY}": self.diagnosis_secondary,
            "{ICUDATE}": self.icu_date,
            "{CREATIONDATE}": self.icu_date,
            "{NEXTDAY}": (datetime.today() + timedelta(days=1)).strftime("%d.%m.%Y"),
            "{DOCTOR}": self.doctor,
            "{MASS}": str(randint(75, 100)),
            "{HEIGHT}": str(randint(165, 179)),
            "{BMI}": str(randint(75, 100) // (uniform(1.65, 1.79) ** 2))
        }

    def word_bed(self):
        """
        replaces keywords with patients data in a template file, and saves it as a new file
        :return: None
        """
        replacements = [(key, value) for key, value in self.create_replacements().items()]

        replace_docx_text(
            os.path.normpath("TemplateFiles/nakrovatnik.docx"),
            os.path.normpath(f"{self.path}/{self.name.split()[0]}_нарковатник.docx"),
            ("{NAME}", self.name),
            *replacements,
            html=True
        )

    def prescription_list(self):
        """
        creates a prescription Excel list from a template list
        :return: None
        """
        replacements = self.create_replacements()

        workbook = openpyxl.load_workbook(os.path.normpath("TemplateFiles/karta.xlsx"))

        worksheet = workbook["."]

        if not os.path.exists(os.path.normpath("TemplateFiles/excel_map.txt")):
            # call mapper method only at a 1st run to create a mapping dictionary of key_word: coordinate
            self.mapper(worksheet=worksheet, replacements=replacements)

        with open(file=os.path.normpath("TemplateFiles/excel_map.txt"), mode='r') as file:
            # get key_word: coordinate dict from excel_map.txt file
            key_coordinate = json.load(file)

        for k, v in key_coordinate.items():
            str_to_change = worksheet.cell(*coordinate_to_tuple(v)).value
            worksheet.cell(*coordinate_to_tuple(v)).value = str_to_change.replace(k, replacements[k])

        workbook.save(os.path.normpath(f"{self.path}/{self.name.split()[0]}.xlsx"))

    def worksheet_update(self):
        """
        Creates addition data list to fill gsheets with and pass it to actual filling method from utils.GSheet
        :return: None
        """
        data = [
            self.name,
            self.age,
            self.hospitalisation_date,
            self.icu_date,
            1,
            self.department,
            None,
            self.diagnosis_main,
            self.operation,
            self.diagnosis_secondary,
        ]

        spreadsheet_fill(data=data)

    @staticmethod
    def mapper(worksheet: openpyxl.workbook, replacements: dict):
        """
        Creates a word_to_replace: cell_coordinate dict that is dumped to txt to speed up changing Excel file
        :param worksheet: worksheet to iterate over
        :param replacements: dict of key_words: replacement_values
        :return:
        """
        key_coordinate = {}
        for row in worksheet.iter_rows():
            for cell in row:
                v = str(cell.value)
                for key, value in replacements.items():
                    if v is not None and key in v:
                        key_coordinate[key] = cell.coordinate

        key_coordinate = json.dumps(key_coordinate, indent=4)

        with open(file=os.path.normpath("TemplateFiles/excel_map.txt"), mode="w") as file:
            file.write(key_coordinate)


if __name__ == '__main__':
    pass
